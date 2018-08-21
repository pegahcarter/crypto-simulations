import pandas as pd
import numpy as np
import ccxt
from datetime import datetime
from openpyxl import load_workbook
from xlsxwriter.utility import xl_rowcol_to_cell as cell

#8/21/18

def update_quantity(add_side, add_amt, subtract_side, subtract_amt):
    data[1][data[0].index(add_side)] += add_amt
    data[1][data[0].index(subtract_side)] -= subtract_amt


def rebalance(coin1, coin2, weight_to_sell, tickers):
    if coin1 + '/' + coin2 in tickers:
        order = coin1 + '/' + coin2, 'sell'
    elif coin2 + '/' + coin1 in tickers:
        order = coin2 + '/' + coin1
    else:
        order = coin1 + '/BTC', 'sell', coin2 + '/BTC', 'buy'

    dollar_amt = abs(weight_to_sell * data['dollar_value'].sum())
    for i in range(0, order, 2):
        coin_amt = float(dollar_amt / data.loc[data['symbol'] == order[i][:3], 'price'])
        single_trade = 'Y'
        if len(order) > 2:
            single_trade = 'N'

        trade_coin(order[i], order[i+1], coin_amt, dollar_amt, fees, single_trade)


def trade_coin(ratio, side, coin_amt, dollar_amt, single_trade):
    exchange.create_order(ratio, 'market', side, coin_amt)
    transactions = pd.read_excel('transactions.xlsx')
    num_rows = len(transactions.index)
    trade_id = transactions['trade_id'][num_rows] + 1
    trade_date = datetime.now()
    fees = dollar_amt * .00075
    transaction = [trade_id, trade_date, side, ratio[:3], ratio[4:], coin_amt, dollar_amt, fees, single_trade]
    for i in range(transaction):
        wb = load_workbook('transactions.xlsx')
        ws = wb.worksheets[0]
        ws.cell(row=num_row+1, column=i+1).value = transaction[i]
        wb.save('transactions.xlsx')


file = 'C:/Users/Carter Carlson/Documents/Administrative/api.txt'
with open(file, 'r') as f:
    api = f.readlines()
    apiKey = api[0][:len(api[0])-1]
    secret = api[1][:len(api[1])]

exchange = ccxt.binance({
    'options': {'adjustForTimeDifference': True},
    'apiKey': apiKey,
    'secret': secret})

balance = exchange.fetchBalance()
wallet = pd.DataFrame(balance['info']['balances'])
data, price, quantity, coins = [], [], [], []
tickers = set()

[tickers.add(ticker) for ticker in exchange.fetchTickers()]
coins = wallet.loc[wallet['free'].astype(float) > .1, 'asset'].tolist()
data = update_data(coins)
thresh = .01
num_coins = len(data)
avg_weight = 1/num_coins
weight_thresh = avg_weight * thresh



btc_price = float(exchange.fetch_ticker('BTC/USDT')['info']['lastPrice'])
quantity = [balance[coin]['total'] for coin in coins]
price.append([float(exchange.fetch_ticker(coin + '/BTC')['info']['lastPrice']) * btc_ratio for coin in coins if coin != 'BTC' else btc_price])
data.append([coin_list, quantity, price, quantity*price], np.divide(data['dollar_value'], sum(price)))

while True:
    light_weight, heavy_weight = min(data[2]), max(data[2])
    light_coin, heavy_coin = data[0][data[2].index(light_weight)], data[0][data[2].index(heavy_weight)]

    if heavy_weight - light_weight < 2 * avg_weight * thresh:
        break
    elif avg_weight - light_weight < heavy_weight - avg_weight:
        weight_to_sell = (heavy_weight - avg_weight)
    else:
        weight_to_sell = (avg_weight - light_weight)

    dollar_amt = weight_to_sell * total_dollar_value
    ratios = [light_coin + '/' + heavy_coin, heavy_coin + '/' + light_coin]
    ticker = list((set(ratios) & tickers))

    if len(ticker) < 1:
        ticker = [heavy_coin + '/BTC', light_coin + '/BTC']
        denom = ticker[1][:ticker[1].find('/')]
        trade_count += 2
    else:
        denom = ticker[0][ticker[0].find('/') + 1:]
        trade_count += 1

    numer = ticker[0][:ticker[0].find('/')]
    numer_quantity = np.divide(dollar_amt, small_historical_prices[num_day][data[0].index(numer)])
    denom_quantity = np.divide(dollar_amt, small_historical_prices[num_day][data[0].index(denom)])
    denom_quantity_after_fees = (1-rate) * denom_quantity

    if numer == light_coin:
        update_quantity(numer, numer_quantity, denom, denom_quantity_after_fees)
    else:
        update_quantity(denom, denom_quantity_after_fees, numer, numer_quantity)






    rebalance(data['symbol'][0], data['symbol'][num_coins-1], weight_to_sell, tickers)
    data = update_data(coins)


----------------------------------------------------------------------------
# 8/14/18
def update_data(coins):
    df = pd.DataFrame(columns=['symbol', 'quantity', 'price', 'dollar_value'])
    btc_price = float(exchange.fetch_ticker('BTC/USDT')['info']['lastPrice'])
    for coin in coins:
        quantity = balance[coin]['total']
        price = btc_price
        if coin != 'BTC':
            btc_ratio = float(exchange.fetch_ticker(coin + '/BTC')['info']['lastPrice'])
            price *= btc_ratio

        dollar_value = quantity * price
        df = df.append({'symbol': coin,'quantity':quantity,'price':price,'dollar_value':dollar_value}, ignore_index=True)

    df = df.sort_values('dollar_value', ascending=False).reset_index(drop=True)
    df['weight'] = df['dollar_value'].divide(data['dollar_value'].sum())
    return df


def rebalance(coin1, coin2, weight_to_sell, tickers):
    if coin1 + '/' + coin2 in tickers:
        order = coin1 + '/' + coin2, 'sell'
    elif coin2 + '/' + coin1 in tickers:
        order = coin2 + '/' + coin1
    else:
        order = coin1 + '/BTC', 'sell', coin2 + '/BTC', 'buy'

    dollar_amt = abs(weight_to_sell * data['dollar_value'].sum())
    for i in range(0, order, 2):
        coin_amt = float(dollar_amt / data.loc[data['symbol'] == order[i][:3], 'price'])
        single_trade = 'Y'
        if len(order) > 2:
            single_trade = 'N'

        trade_coin(order[i], order[i+1], coin_amt, dollar_amt, fees, single_trade)


def trade_coin(ratio, side, coin_amt, dollar_amt, single_trade):
    exchange.create_order(ratio, 'market', side, coin_amt)
    transactions = pd.read_excel('transactions.xlsx')
    num_rows = len(transactions.index)
    trade_id = transactions['trade_id'][num_rows] + 1
    trade_date = datetime.now()
    fees = dollar_amt * .00075
    transaction = [trade_id, trade_date, side, ratio[:3], ratio[4:], coin_amt, dollar_amt, fees, single_trade]
    for i in range(transaction):
        wb = load_workbook('transactions.xlsx')
        ws = wb.worksheets[0]
        ws.cell(row=num_row+1, column=i+1).value = transaction[i]
        wb.save('transactions.xlsx')


file = 'C:/Users/Carter Carlson/Documents/Administrative/api.txt'
with open(file, 'r') as f:
    api = f.readlines()
    apiKey = api[0][:len(api[0])-1]
    secret = api[1][:len(api[1])]

exchange = ccxt.binance({
    'options': {'adjustForTimeDifference': True},
    'apiKey': apiKey,
    'secret': secret})

balance = exchange.fetchBalance()
wallet = pd.DataFrame(balance['info']['balances'])

coins = []
tickers = set()

[tickers.add(ticker) for ticker in exchange.fetchTickers()]
coins = wallet.loc[wallet['free'].astype(float) > .1, 'asset'].tolist()
data = update_data(coins)
thresh = .01
num_coins = len(data)
avg_weight = 1/num_coins
weight_thresh = avg_weight * thresh

while True:
    heavy_weight_dif = data['weight'][0] - avg_weight
    light_weight_dif = avg_weight - data['weight'][num_coins-1]
    if heavy_weight_dif < weight_thresh and light_weight_dif < weight_thresh:
        break
    elif heavy_weight_dif > light_weight_dif:
        weight_to_sell = light_weight_dif
    else:
        weight_to_sell = heavy_weight_dif

    rebalance(data['symbol'][0], data['symbol'][num_coins-1], weight_to_sell, tickers)
    data = update_data(coins)

--------------------------------------------------------------------------------
# 8/1/18
def update_data(coins):
    df = pd.DataFrame(columns=['symbol', 'quantity', 'price', 'dollar_value'])
    btc_price = float(exchange.fetch_ticker('BTC/USDT')['info']['lastPrice'])
    for coin in coins:
        quantity = balance[coin]['total']
        price = btc_price
        if coin != 'BTC':
            btc_ratio = float(exchange.fetch_ticker(coin + '/BTC')['info']['lastPrice'])
            price *= btc_ratio

        dollar_value = quantity * price
        df = df.append({'symbol': coin,'quantity':quantity,'price':price,'dollar_value':dollar_value}, ignore_index=True)

    df = df.sort_values('dollar_value', ascending=False).reset_index(drop=True)
    df['weight'] = df['dollar_value'].divide(data['dollar_value'].sum())
    return df


def write_to_excel(ratio, side, coin_amt, dollar_amt, fees):
    wb = load_workbook(transaction_file)
    sheet = wb.active
    rebalance_id = sheet.cell(row=sheet.max_row, column=2).value + 1
    row_num = sheet.max_row
    trade_id = sheet.cell(row=row_num, column=1).value + 1
    trade_date = datetime.now()
    symbol1 = ratio[:3]
    symbol2 = ratio[4:]
    single_trade = 'Y'
    if no_ticker:
        single_trade = 'N'

    transaction = [trade_id, rebalance_id, trade_date, side, symbol1, symbol2, coin_amt, dollar_amt, fees, single_trade]
    for i in range(10):
        sheet.cell(row=row_num, column=i+1).value = transaction[i]

    wb.save(file)


def get_ticker(coin1, coin2):
    if coin1 + '/' + coin2 in tickers:
        return coin1 + '/' + coin2,
    elif coin2 + '/' + coin1 in tickers:
        return coin2 + '/' + coin1,
    else:
        return coin1 + '/BTC', coin2 + '/BTC'


def trade_coin(ratio, side, coin_amt, dollar_amt, fees):
    exchange.create_order(ratio, 'market', side, coin_amt, param)
    write_to_excel(ratio, side, coin_amt, dollar_amt, fees)


def rebalance(coin1, coin2):
    dollar_amt = abs(weight_to_sell *  data['dollar_value'].sum())
    fees = dollar_amt * .00075
    ticker = get_ticker(coin1, coin2)
    side = 'sell'
    try:
        coin_amt = float(dollar_amt / data.loc[data['symbol'] == ticker[1][:3], 'price'])
        no_ticker = True
        trade_coin(ticker[1], side, coin_amt, dollar_amt, fees)
    except:
        pass
    if any(coin2 + '/' in x for x in ticker):
        side = 'buy'

    coin_amt = float(dollar_amt / data.loc[data['symbol'] == ticker[0][:3], 'price'])
    trade_coin(ticker[0], side, coin_amt, dollar_amt, fees)


api_file = "C:/Users/Carter Carlson/Documents/Excel References/secret.csv"
transaction_file = 'transactions.xlsx'
no_ticker = None
api = pd.read_csv(api_file)
exchange = ccxt.binance({'options': {'adjustForTimeDifference': True},'apiKey': api['apiKey'][0],'secret': api['secret'][0]})
balance = exchange.fetchBalance()
wallet = pd.DataFrame(balance['info']['balances'])

coins = []
heavy_coins = []
light_coins = []
tickers = set()
[tickers.add(ticker) for ticker in exchange.fetchTickers()]
coins = wallet.loc[wallet['free'].astype(float) > .1, 'asset'].tolist()
data = update_data(coins)
thresh = .01
num_coins = len(data)
avg_weight = 1/num_coins
weight_thresh = avg_weight * thresh

while True:
    heavy_weight_dif = data['weight'][0] - avg_weight
    light_weight_dif = avg_weight - data['weight'][num_coins-1]
    if heavy_weight_dif < weight_thresh and light_weight_dif < weight_thresh:
        print('a')
        break
    elif heavy_weight_dif > light_weight_dif:
        print('b')
        weight_to_sell = light_weight_dif
    else:
        print('c')
        weight_to_sell = heavy_weight_dif

    rebalance(data['symbol'][0], data['symbol'][num_coins-1])
    data = update_data(coins)


--------------------------------------------------------------------------------------------------
# 7/25/18

def write_to_excel(ticker, side, coin_amt):

    row_num = sheet.max_row
    trade_id = sheet.cell(row=row_num, column=1).value + 1
    trade_date = datetime.now()
    symbol1 = ticker[:3]
    symbol2 = ticker[4:]
    fees = trade_in_dollars * .00075
    transaction = [trade_id, rebalance_id, trade_date, side, symbol1, symbol2, amt, trade_in_dollars, fees, single_trade]
    [sheet.cell(row=row_num, column=i+1).value = transaction[i] for i in range(10)]
    wb.save(file)


def update_data(coins):

    df = pd.DataFrame(columns=['symbol', 'quantity', 'price', 'dollar_value'])
    btc_price = float(exchange.fetch_ticker('BTC/USDT')['info']['lastPrice'])
    for coin in coins:
        quantity = balance[coin]['total']
        price = btc_price
        if coin != 'BTC':
            btc_ratio = float(exchange.fetch_ticker(coin + '/BTC')['info']['lastprice'])
            price *= btc_ratio

        dollar_value = quantity * price
        df = df.append({'symbol': coin,'quantity':quantity,'price':price,'dollar_value':dollar_value}, ignore_index=True)

    df = df.sort_values('dollar_value', ascending=False).reset_index()
    df['weight'] = df['dollar_value'].divide(df['dollar_value'].sum())
    return df


def rebalance_order(coin1, coin2, coin2_weight_dif):

    trade_in_dollars = coin2_weight_dif * port_dollar_value
    amt = trade_in_dollars / light_value # Note: is light_value correct? do I need to change?
    single_trade = 'Y'
    try:
        side = 'buy'
        exchange.fetch_ticker(coin2 + '/' + coin1)['info']
        ticker = coin2 + '/' + coin1
    except:
        try:
            side = 'sell'
            exchange.fetch_ticker(coin1 + '/' + coin2)['info']
            ticker =  coin1 + '/' + coin2
        except:
            single_trade = 'N'
            ticker = coin1 + '/BTC'
            print(exchange.create_order(ticker, 'market', side, amt, param))
            write_to_excel()
            side = 'buy'
            ticker = coin2 + '/BTC'

    finally:
        print(exchange.create_order(ticker, 'market', side, amt, param))
        write_to_excel()
        data.loc[data['weight'] == coin1] -= coin2_weight_dif
        data.loc[data['weight'] == coin2] += coin2_weight_dif


def get_coin_info(coin):

    dollar_value, weight = data.loc[coin, ['dollar_value', 'weight']].tolist()
    weight_dif = abs(weight - avg_weight)
    return coin, dollar_value, weight, weight_dif


api_file = "C:/Users/Carter Carlson/Documents/Excel References/secret.csv"
transaction_file = 'transactions.xlsx'

api = pd.read_csv(api_file)
exchange = ccxt.binance({'options': {'adjustForTimeDifference': True},'apiKey': api['apiKey'][0],'secret': api['secret'][0]})
balance = exchange.fetchBalance()
wallet = balance['info']['balances']

coins = []
heavy_coins = []
light_coins = []
coins = wallet.loc[wallet['free'].astype(float) > .1, 'asset'].tolist()

data = update_data(coins)
avg_weight = len(data)
heavy_coins = data.loc[data['weight'] > avg_weight, 'symbol'].tolist()
light_coins = data.loc[~data['symbol'].isin(heavy_coins), 'symbol'].tolist()

wb = load_workbook(transaction_file)
sheet = wb.active
rebalance_id = sheet.cell(row=sheet.max_row, column=2).value + 1

param = {'test':True}
avg_weight = 1/len(coins)
thresh = .05

for a in range(len(heavy_coins)):
    for b in range(len(light_coins)):
        heavy_coin, heavy_value, heavy_weight, heavy_weight_dif = get_coin_info(heavy_coins[a])
        light_coin, light_value, light_weight, light_weight_dif = get_coin_info(light_coins[b])
        if abs(heavy_weight_dif - light_weight_dif) <= 2 * thresh * avg_weight:
            break
        elif heavy_weight_dif > light_weight_dif:
            rebalance_order(heavy_coin, light_coin, light_weight_dif)
            break
        else:
            for c in range(a + 1, len(heavy_coins)):
                heavy_coin, heavy_value, heavy_weight, heavy_weight_dif = get_coin_info(heavy_coins[c])
                if abs(heavy_weight_dif - light_weight_dif) <= 2 * thresh * avg_weight:
                    break
                elif light_weight_dif > heavy_weight_dif:
                    rebalance_order(light_coin, heavy_coin, heavy_weight_dif)
                else:
                    rebalance_order(heavy_coin, light_coin, light_weight_dif)
                    break
