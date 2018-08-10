import pandas as pd
import numpy as np
import ccxt
from datetime import datetime
from openpyxl import load_workbook

def write_to_excel(ticker, side, coin_amt):

    row_num = sheet.max_row
    trade_id = sheet.cell(row=row_num, column=1).value + 1
    trade_date = datetime.now()
    symbol1 = ticker[:3]
    symbol2 = ticker[4:]
    fees = trade_in_dollars * .00075
    transaction = [trade_id, rebalance_id, trade_date, side, symbol1, symbol2, amt, trade_in_dollars, fees, single_trade]
    [sheet.cell(row=row_num, column=i+1).value = transaction[i] for i in range(10)]
    wb.save(file)


def update_data(coins):

    df = pd.DataFrame(columns=['symbol', 'quantity', 'price', 'dollar_value'])
    btc_price = float(exchange.fetch_ticker('BTC/USDT')['info']['lastPrice'])
    for coin in coins:
        quantity = balance[coin]['total']
        price = btc_price
        if coin != 'BTC':
            btc_ratio = float(exchange.fetch_ticker(coin + '/BTC/')['info']['lastprice'])
            price *= btc_ratio

        dollar_value = quantity * price
        df = df.append({'symbol': coin,'quantity':quantity,'price':price,'dollar_value':dollar_value}, ignore_index=True)

    df = df.sort_values('dollar_value', ascending=False)
    return df


def rebalance_order(coin1, coin2, coin2_weight_dif):

    trade_in_dollars = coin2_weight_dif * port_dollar_value
    amt = trade_in_dollars / light_value # Note: is light_value correct? do I need to change?
    single_trade = 'Y'
    try:
        side = 'buy'
        exchange.fetch_ticker(coin2 + '/' + coin1)['info']
        ticker = coin2 + '/' + coin1
    except:
        try:
            side = 'sell'
            exchange.fetch_ticker(coin1 + '/' + coin2)['info']
            ticker =  coin1 + '/' + coin2
        except:
            single_trade = 'N'
            ticker = coin1 + '/BTC'
            print(exchange.create_order(ticker, 'market', side, amt, param))
            write_to_excel()
            side = 'buy'
            ticker = coin2 + '/BTC'

    finally:
        print(exchange.create_order(ticker, 'market', side, amt, param))
        write_to_excel()
        data.loc[data['weight'] == coin1] -= coin2_weight_dif
        data.loc[data['weight'] == coin2] += coin2_weight_dif


def get_coin_info(coin):

    dollar_value, weight = data.loc[coin, ['dollar_value', 'weight']].tolist()
    weight_dif = abs(weight - avg_weight)
    return coin, dollar_value, weight, weight_dif


api_file = "C:/Users/Carter Carlson/Documents/Excel References/secret.csv"
transaction_file = 'transactions.xlsx'

api = pd.read_csv(api_file)
exchange = ccxt.binance({'options': {'adjustForTimeDifference': True},'apiKey': api['apiKey'][0],'secret': api['secret'][0]})
balance = exchange.fetchBalance()
wallet = balance['info']['balances']

coins = []
heavy_coins = []
light_coins = []
coins = wallet.loc[wallet['free'].astype(float) > .1, 'asset'].tolist()

data = update_data(coins)
port_dollar_value = data['dollar_value'].sum()
data['weight'] = data['dollar_value'].divide(port_dollar_value)
heavy_coins = data.loc[data['weight'] > avg_weight, 'symbol'].tolist()
light_coins = data.loc[~data['symbol'].isin(heavy_coins), 'symbol'].tolist()

wb = wb.load_workbook(transaction_file)
sheet = wb.active
rebalance_id = sheet.cell(row=sheet.max_row, column=2).value + 1

param = {'test':True}
avg_weight = 1/len(coins)
thresh = .05

for a in range(len(heavy_coins)):
    for b in range(len(light_coins)):
        heavy_coin, heavy_value, heavy_weight, heavy_weight_dif = get_coin_info(heavy_coins[a])
        light_coin, light_value, light_weight, light_weight_dif = get_coin_info(light_coins[b])
        if abs(heavy_weight_dif - light_weight_dif) <= 2 * thresh * avg_weight:
            break
        elif heavy_weight_dif > light_weight_dif:
            rebalance_order(heavy_coin, light_coin, light_weight_dif)
            break
        else:
            for c in range(a + 1, len(heavy_coins)):
                heavy_coin, heavy_value, heavy_weight, heavy_weight_dif = get_coin_info(heavy_coins[c])
                if abs(heavy_weight_dif - light_weight_dif) <= 2 * thresh * avg_weight:
                    break
                elif light_weight_dif > heavy_weight_dif:
                    rebalance_order(light_coin, heavy_coin, heavy_weight_dif)
                else:
                    rebalance_order(heavy_coin, light_coin, light_weight_dif)
                    break
--------------------------------------------------------------------------------
# Testing

def write_to_excel():
    row_num = sheet.max_row
    trade_id = sheet.cell(row=row_num, column=1).value + 1
    trade_date = datetime.now()
    symbol1 = ratio[:3]
    symbol2 = ratio[4:]
    fees = dollar_amt * .00075
    transaction = [trade_id, rebalance_id, trade_date, side, symbol1, symbol2, coin_amt, dollar_amt, fees, single_trade]
    [sheet.cell(row=row_num, column=i+1).value = transaction[i] for i in range(10)]
    wb.save(file)


def get_ratio(coin1, coin2):
    try:
        exchange.fetch_ticker(coin1 + '/' + coin2)['info']
        return coin1 + '/' + coin2, True
    except:
        try:
            exchange.fetch_ticker(coin2 + '/' + coin1)['info']
            return coin2 + '/' + coin1, True
        except:
            return coin1 + '/BTC', False


def trade_coin():
    exchange.create_order(ratio, 'market', side, coin_amt, param)
    write_to_excel()


def rebalance(coin1, coin2): #larger coin, smaller coin
    dollar_amt = trade_in_dollars(coin1)
    ratio, single_trade = get_ratio(coin1, coin2)
    side = 'sell'
    if coin2 = ratio[:3]:
        side = 'buy'

    if not single_trade:
        coin_amt = get_coin_amt(dollar_amt, coin1)
        trade_coin()
        ratio = coin2 + '/BTC'
        side = 'sell'

    coin_amt = get_coin_amt(dollar_amt, coin2)
    trade_coin()



def get_coin_amt(dollar_amt, coin):
    return dollar_amt / data.loc[data['symbol'] == coin, 'price']


def trade_in_dollars(coin):
    weight_dif = abs(data.loc[data['symbol'] == coin, 'weight'] - avg_weight)
    return weight_dif * port_dollar_value










































print()
